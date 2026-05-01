from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.tokens import default_token_generator, PasswordResetTokenGenerator
from django.contrib.auth.hashers import make_password
from django.contrib.auth.decorators import login_required
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.shortcuts import render, redirect, get_object_or_404
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.cache import cache
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import os
from .forms import SignupForm, GSTDetailsForm, UserManagementForm

# For E-Invoice Comparison View
# Change 1 for E-Invoice Register feature
from .models import CompanyDetails, Invoice, GSTDetails, HSN, SAC, InvoiceRemark, EInvoiceRegister

import random 
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib.sites.shortcuts import get_current_site
from decimal import Decimal, InvalidOperation
from datetime import datetime
from .tokens import custom_token_generator
from .utils import send_email
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Invoice
import json
import re
from accounts.validations.invoice_mapper import run_dynamic_validations
from accounts.validations.invoice_mapper import call_ocr_api, map_api_data_to_invoice

# Step1 : before validation we must populate invoice model
from accounts.validations.data_gathering import api_response_test
from accounts.validations.invoice_mapper import map_api_data_to_invoice


# ====== Additional imports for data processing for modal popup for Address Matching and Invoice Calculation ===== #
#import re 
#from difflib import SequenceMatcher


User = get_user_model()


def normalize_user_role(role):
    """Map legacy role values to current role constants."""
    role_aliases = {
        'SuperUser': 'COMPANY_ADMIN',
        'Processor': 'PROCESSOR',
    }
    return role_aliases.get(role, role)



# Helper function to normalize and compare addresses for better matching in Account Check (Customer Address)
'''
def normalize_address(text):
    if not text:
        return ""
    
    text = text.lower()
    text = re.sub(r'[^a-z0-9 ]', ' ', text) # Remove special characters
    text = re.sub(r'\s+', ' ', text)  # normalize spaces

    # remove meaningless words
    stop_words = {
        "road", "rd", "street", "st", "area", "village",
        "district", "state", "india", "none"
    }

    tokens = [w for w in text.split() if w not in stop_words]
    return " ".join(tokens)
'''

'''
# Helper function to calculate address match percentage for Account Check (Customer Address)
def address_match_percentage(addr1, addr2):
    addr1_norm = normalize_address(addr1)
    addr2_norm = normalize_address(addr2)

    if not addr1_norm or not addr2_norm:
        return 0
    
    return SequenceMatcher(None, addr1_norm, addr2_norm).ratio() * 100
'''   

#----------------------------------------------------------------------------------------------------------------------------------------

# Signup View with email verification and password setup link
@csrf_protect
def signup(request):
    if request.method == 'POST':
        form = SignupForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    company = form.save()


                    # Create the User entry
                    user = User.objects.create(
                        username=company.contact_person_name,
                        email=company.contact_person_email,
                        company_code=company,
                        role='COMPANY_ADMIN',
                    )
                    user.set_unusable_password()
                    user.save()
                    

                    # Send Reset Password Link
                    token = custom_token_generator.make_token(user)
                    uid = urlsafe_base64_encode(user.email.encode('utf-8'))
                    domain = request.get_host()
                    link = f"http://{domain}/reset-password/{uid}/{token}/"
                    
                    subject = "Set your password"
                    body = f'''
                        <html>
                            <body>
                                <p>Dear {company.contact_person_name},</p>
                                <p>Click the link below to reset your password:</p>
                                <p><a href="{link}" target="_blank">Reset Password</a></p>
                            </body>
                        </html>
                    '''
                    
                    # Send using SMTP with configured credentials from settings
                    send_email(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD,
                               company.contact_person_email, subject, body=body)

                    messages.success(request, 'Account created successfully! Check your email to set your password.')
                    return redirect('password_reset_sent')


            except Exception as e:
                print(f"Signup process failed: {str(e)}")
                messages.error(request, 'Signup failed. Please try again.')
                return render(request, 'signup.html', {'form': form})
        else:
            return render(request, 'signup.html', {'form': form})

    # Initial GET request → show signup form
    form = SignupForm()
    return render(request, 'signup.html', {'form': form})


# Password Reset View
@csrf_protect
def reset_password(request, uidb64, token):

    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        print(f"DEBUG: Decoded UID: {uid}")
        user = User.objects.get(email=uid)
        print(f"DEBUG: User found: {user.email}")
    except (TypeError, ValueError, OverflowError, User.DoesNotExist) as e:
        print(f"DEBUG: Error finding user: {e}")
        user = None

    if user:
        token_valid = custom_token_generator.check_token(user, token)
        print(f"DEBUG: Token valid: {token_valid}")
        if token_valid:
            if request.method == 'POST':
                new_password = request.POST.get('password')
                confirm_password = request.POST.get('confirm_password')
                
                # here we are checking for password match so on UI also there will be 2 fields 
                # 1st is password and 2nd is confirm password 
                if new_password == confirm_password:
                    user.set_password(new_password)
                    user.save()
                    messages.success(request, 'Password set successfully! You can now log in.')
                    return redirect('login')
                else:
                    messages.error(request, 'Passwords do not match.')
                    return render(request, 'reset_password.html', {'uid': uid, 'token': token})
            return render(request, 'reset_password.html', {'uid': uid, 'token': token})
    
    print(f"DEBUG: Token validation failed - user: {user}, token passed check: {token_valid if user else False}")
    messages.error(request, 'The token is expired or invalid.')
    return render(request, 'password_reset_failed.html', {'error': 'The token is expired or invalid.'})



# Login View or Login Functionality 
@csrf_protect
def loginview(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        # Get user by email, then authenticate with username
        try:
            user_obj = User.objects.get(email=email)
        except User.DoesNotExist:
            user_obj = None
        
        if user_obj:
            user = authenticate(request, username=user_obj.username, password=password)
        else:
            user = None
        
        if user is not None:
            role = normalize_user_role(user.role)

            # Force admin users through Django admin login page instead of auto-signing
            # them in via the custom application login flow.
            if user.is_superuser or user.is_staff or role == 'APP_ADMIN':
                logout(request)
                return redirect('/admin/login/?next=/admin/')

            login(request, user)
            request.session["company_code"] = user.company_code_id
            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)

            if role == 'COMPANY_ADMIN':
                return redirect('company_admin_dashboard')
            else:
                return redirect('user_dashboard')
        else:
            error_message = "Entered credentials are incorrect. Please enter correct credentials."
            return render(request, 'login.html', {'error': error_message})
         
    return render(request, 'login.html')


# logout View
@login_required
def logoutview(request):
    logout(request)
    return redirect('login')


# Password Reset Sent View
def password_reset_sent(request):
    return render(request, "password_reset_sent.html")


# Password Reset Done View
def password_reset_done(request):
    return render(request, "password_reset_done.html")


# Added this dashboard but currently not using it as using django admin for app admin for better
# security and less development time but keeping this code ready for future if we want to build

'''
from django.http import HttpResponseForbidden
@login_required
def app_admin_dashboard(request):
    if request.user.role != 'APP_ADMIN':
        return HttpResponseForbidden("Not allowed")

    total_companies = CompanyDetails.objects.count()
    total_users = User.objects.count()

    return render(request, 'app_admin_dashboard.html', {
        'total_companies': total_companies,
        'total_users': total_users,
    })
'''

# SuperUser Dashboard View
from django.http import HttpResponseForbidden
@login_required
def superuser_dashboard(request):
    """
    SuperUser Dashboard handles:
    1. User Management (Create users)
    2. GST Details Management (Add/Edit GST registrations)
    """

    role = normalize_user_role(getattr(request.user, 'role', ''))
    if role != 'COMPANY_ADMIN':
        return HttpResponseForbidden("Not allowed")
    
    # Handle POST requests for both User Creation and GST Details
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # USER CREATION
        if action == 'create_user':
            user_form = UserManagementForm(request.POST)
            if user_form.is_valid():
                try:
                    user = user_form.save(commit=False, company=request.user.company_code)
                    user.save()
                    messages.success(request, f'User {user.username} created successfully!')
                    response = redirect('company_admin_dashboard')
                    response['Location'] = response['Location'] + '?section=users'
                    return response
                except Exception as e:
                    messages.error(request, f'Failed to create user: {str(e)}')
            else:
                # Form validation errors
                for field, errors in user_form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
        
        # GST DETAILS CREATION
        elif action == 'add_gst':
            gst_form = GSTDetailsForm(request.POST)
            if gst_form.is_valid():
                try:
                    gst_detail = gst_form.save(commit=False)
                    gst_detail.company = request.user.company_code
                    gst_detail.created_by = request.user
                    gst_detail.save()
                    messages.success(request, f'GST Details for {gst_detail.state} added successfully!')
                    response = redirect('company_admin_dashboard')
                    response['Location'] = response['Location'] + '?section=settings'
                    return response
                except Exception as e:
                    messages.error(request, f'Failed to add GST details: {str(e)}')
            else:
                # Form validation errors
                for field, errors in gst_form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
    

    # Fetch all users and GST records from the same company using stable id lookup.
    company_id = getattr(request.user, 'company_code_id', None)
    if not company_id:
        session_company_id = request.session.get('company_code')
        try:
            company_id = int(session_company_id) if session_company_id else None
        except (TypeError, ValueError):
            company_id = None

    users = User.objects.none()
    gst_details = GSTDetails.objects.none()
    if company_id:
        users = User.objects.filter(company_code_id=company_id).order_by('-date_joined')
        gst_details = GSTDetails.objects.filter(company_id=company_id).order_by('state')
    

    # Create empty forms for the template
    user_form = UserManagementForm()
    gst_form = GSTDetailsForm()
    

    context = {
        #'user_role': getattr(request.user, 'role', 'SuperUser'),
        'user_role': role,

        'company': getattr(request.user, 'company_code', None),
        'users': users,
        'gst_details': gst_details,
        'user_form': user_form,
        'gst_form': gst_form,
    }

    #return render(request, 'superuser_dashboard.html', context)
    return render(request, 'company_admin_dashboard.html', context) #



# User Dashboard View with sections
@login_required
def user_dashboard(request):
    """
    User Dashboard with sections: Upload invoices , Pending Action (Standing status) , On Hold (Hold status) , Rejected (Rejected status),
    Approved (Approved status) , Accounted (Accounted status)
    
    Per-column prefix filters + pagination per section
    """
    
    # Current active section
    section = request.GET.get('section', 'upload')

    # Base queryset (company-level isolation)
    # Resolve company from stable id first, then keep session in sync.
    company_obj = None
    company_id = getattr(request.user, "company_code_id", None)

    if not company_id:
        session_company_id = request.session.get("company_code")
        try:
            company_id = int(session_company_id) if session_company_id else None
        except (TypeError, ValueError):
            company_id = None

    if company_id:
        company_obj = CompanyDetails.objects.filter(pk=company_id).first()
        if company_obj and request.session.get("company_code") != company_obj.pk:
            request.session["company_code"] = company_obj.pk

    invoices_qs = Invoice.objects.none()
    if company_obj is not None:
        invoices_qs = Invoice.objects.filter(
            company=company_obj
        ).order_by('-upload_date')

    
    #-------------------------------------------------------------------------------------------
    # Summary Counts for dashboard cards 
    # we are doing separate counts for each status to show on dashboard
    # invoices_qs is base queryset for all invoices of the company
    pending_count = invoices_qs.filter(Q(status__iexact='Standing') | Q(status__iexact='Pending')).count()
    hold_count = invoices_qs.filter(status__iexact='Hold').count()
    rejected_count = invoices_qs.filter(status__iexact='Rejected').count()
    approved_count = invoices_qs.filter(status__iexact='Approved').count()
    

    # we are not showing accounter count as of now on dashboard as per new BRD but keeping the code ready for future if we need to show it later
    # change 1 for accounted status 
    #accounted_count = invoices_qs.filter(status='Accounted').count()
    #-------------------------------------------------------------------------------------------


    # Get filter values
    vendor_prefix = request.GET.get('vendor_prefix', '')
    invoice_no_prefix = request.GET.get('invoice_no_prefix', '')

    #date_prefix = request.GET.get('date_prefix', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    amount_min = request.GET.get('amount_min', '')
    amount_max = request.GET.get('amount_max', '')
    status_prefix = request.GET.get('status_prefix', '')
    response_prefix = request.GET.get('response_prefix', '')
    file_prefix = request.GET.get('file_prefix', '')

    # Avoid cross-tab stale status filters hiding data on refresh.
    normalized_status_for_section = {
        'pending': {'standing', 'pending'},
        'hold': {'hold'},
        'rejected': {'rejected'},
        'approved': {'approved'},
    }
    allowed_status_values = normalized_status_for_section.get(section)
    if allowed_status_values and status_prefix and status_prefix.strip().lower() not in allowed_status_values:
        status_prefix = ''


    # Invoice Upload Limit Logic for USAGE METER CARD on dashboard - max 50 invoices per company as per BRD and show remaining usage
    
    MAX_INVOICES = 50

    # Usage limit is company-wide, not per-uploader.
    user_invoice_count = invoices_qs.count()

    remaining_invoices = max(0, MAX_INVOICES - user_invoice_count)
    usage_percentage = min(100, (user_invoice_count / MAX_INVOICES) * 100)

    
    # Filtering logic
    def apply_filters(qs):

        if vendor_prefix:
            qs = qs.filter(vendor_name__istartswith=vendor_prefix)

        if invoice_no_prefix:
            qs = qs.filter(invoice_number__istartswith=invoice_no_prefix)

        #if date_prefix:
        #    qs = qs.filter(invoice_date__startswith=date_prefix)

        # Date filtering based on from_date and to_date
        if from_date:
            qs = qs.filter(invoice_date__gte=from_date)
        if to_date:
            qs = qs.filter(invoice_date__lte=to_date)

        if status_prefix:
            # Backward compatibility: old UI/query params used "Pending" while DB stores "Standing".
            normalized_status = {
                'pending': 'Standing',
            }.get(status_prefix.strip().lower(), status_prefix)
            qs = qs.filter(status__iexact=normalized_status)

        if response_prefix:
            qs = qs.filter(response__istartswith=response_prefix)

        if file_prefix:
            qs = qs.filter(file_name__istartswith=file_prefix)

        try:
            if amount_min:
                qs = qs.filter(invoice_value__gte=Decimal(amount_min))
            if amount_max:
                qs = qs.filter(invoice_value__lte=Decimal(amount_max))
        except InvalidOperation:
            pass

        return qs


    # 🔽 Helper: Pagination
    def paginate(qs, page_param):
        paginator = Paginator(qs, 10)
        page_number = request.GET.get(page_param, 1)

        try:
            return paginator.page(page_number)
        except PageNotAnInteger:
            return paginator.page(1)
        except EmptyPage:
            return paginator.page(paginator.num_pages)
        

    # 🔽 Section-wise data
    pending_invoices = paginate(
        apply_filters(invoices_qs.filter(Q(status__iexact='Standing') | Q(status__iexact='Pending'))),
        'pending_page'
    )

    hold_invoices = paginate(
        apply_filters(invoices_qs.filter(status__iexact='Hold')),
        'hold_page'
    )

    rejected_invoices = paginate(
        apply_filters(invoices_qs.filter(status__iexact='Rejected')),
        'rejected_page'
    )

    approved_invoices = paginate(
        apply_filters(invoices_qs.filter(status__iexact='Approved')),
        'approved_page'
    )

    
    # change 2 for accounted status
    '''
    accounted_invoices = paginate(
        apply_filters(invoices_qs.filter(status='Accounted')),
        'accounted_page'
    )
    '''

    context = {
        'section': section,

        # Summary counts for dashboard cards 
        'pending_count': pending_count,
        'hold_count': hold_count,
        'rejected_count': rejected_count,
        'approved_count': approved_count,
        #'accounted_count': accounted_count,


        # Paginated invoice lists for each status
        'pending_invoices': pending_invoices,
        'hold_invoices': hold_invoices,
        'rejected_invoices': rejected_invoices,
        'approved_invoices': approved_invoices,
        #'accounted_invoices': accounted_invoices,


        'filters': {
            'vendor_prefix': vendor_prefix,
            'invoice_no_prefix': invoice_no_prefix,

            #'date_prefix': date_prefix,
            'from_date': from_date,
            'to_date': to_date,

            'amount_min': amount_min,
            'amount_max': amount_max,
            'status_prefix': status_prefix,
            'response_prefix': response_prefix,
            'file_prefix': file_prefix,
        },
 
        # Data for Usage Meter Card on Dashboard
        "user_invoice_count": user_invoice_count,
        "remaining_invoices": remaining_invoices,
        "usage_percentage": usage_percentage,
        "max_invoices": MAX_INVOICES,

    }

    return render(request, 'user_dashboard.html', context)


#--------------------------------------------------------------------------------------------------------------------------------------------

# Invoice Views 
# Upload Invoice View
# this view is handling the file upload and also mapping the extracted data to invoice model using the map_api_data_to_invoice function 
@login_required
@csrf_protect
def upload_invoice(request):

    """
    Handle invoice file upload - PDF, Word, JPEG, max 5MB
    """

    if request.method == 'POST':
        files = request.FILES.getlist('invoice_files')
        
        if not files:
            messages.error(request, 'No files selected.')
            return redirect('user_dashboard')
        
        # Allowed extensions per BRD
        allowed_extensions = ['.pdf', '.doc', '.docx', '.jpg', '.jpeg']
        max_size = 5 * 1024 * 1024  # 5MB
        
        uploaded_count = 0
        
        for file in files:
            # Check file extension
            file_ext = os.path.splitext(file.name)[1].lower()
            if file_ext not in allowed_extensions:
                messages.error(request, f'{file.name}: Invalid format. Only PDF, Word, JPEG allowed.')
                continue
            
            
            # Check file size
            if file.size > max_size:
                messages.error(request, f'{file.name}: File exceeds 5MB limit.')
                continue
            

            # Save file
            #fs = FileSystemStorage(location=settings.MEDIA_ROOT / 'invoices')
            fs= FileSystemStorage()
            filename = fs.save(f"invoices/{file.name}", file)
            file_path = fs.url(filename)
            

            # Create Invoice record with status=Standing
            invoice = Invoice.objects.create(
                company=request.user.company_code,
                uploaded_by=request.user,
                vendor_name='Pending',  # To be extracted/manually entered later
                invoice_number='Pending',
                invoice_date=timezone.now().date(),
                invoice_value=0.00,
                file_name=file.name,

                #file_path=str(settings.MEDIA_ROOT / 'invoices' / filename),
                file_path=file_path,

                status='Standing',
                response='Upload successful, awaiting processing'
            )

            uploaded_count += 1
            
            # Extract data using stimulated API response
            #map_api_data_to_invoice(invoice, api_response_test)  
            
            #---------------------------------------------------------------------------
            # Real OCR API call with file path and mapping to invoice model
            #Replace This Line:
            #from validations.invoice_mapper import call_ocr_api


            # Get absolute file path (IMPORTANT)
            absolute_path = fs.path(filename)


            # Registered company PAN from signup form (CompanyDetails.pan)
            # Added company PAN Extraction before API call
            # step 2 : to get company PAN from signup form and pass to API
            company_pan = None
            if getattr(request.user, "company_code", None):
                company_pan = request.user.company_code.pan


            # Call real OCR API 
            print("Company PAN being sent to OCR API:", company_pan)
            # We are passing company PAN to OCR API as well for better accuracy 
            # step 3 : to get company PAN from signup form and pass to API
            ocr_response = call_ocr_api(absolute_path, company_pan=company_pan)
            

            if ocr_response:
                map_api_data_to_invoice(invoice,ocr_response)
                # Keep in Pending Action bucket
                invoice.status = "Standing"
                invoice.response = "OCR and validation completed"
                invoice.save()
            else:
                # Keep visible in Pending Action even on OCR failure
                invoice.status = "Standing"
                invoice.response = "OCR failed (manual review needed)"
                invoice.save()
         
        if uploaded_count > 0:
            messages.success(request, f'{uploaded_count} invoice(s) uploaded successfully.')
        
        return redirect('user_dashboard')
    
    return redirect('user_dashboard')



# 3rd version implementing Invoice button data mapping using session data and popup MODAL
@login_required
@csrf_protect
def update_invoice_status(request, invoice_id):

    if request.method == 'POST':
        new_status = request.POST.get('status')

        if new_status not in ['Approved', 'Rejected', 'Hold']:     # change 3 for accounted status # removed 'Accounted' from list here
            messages.error(request, 'Invalid status.')
            return redirect('user_dashboard')

        try:
            invoice = Invoice.objects.get(
                id=invoice_id,
                company=request.user.company_code
            )

            # STORE RAW DATA ONLY ON APPROVE
            if new_status == 'Approved':
                request.session[f"invoice_raw_data_{invoice.id}"] = api_response_test  # store raw data in session for modal popup

            invoice.status = new_status
            invoice.response = f'Status updated to {new_status} by {request.user.username}'
            invoice.save()

            messages.success(
                request,
                f'Invoice {invoice.invoice_number} marked as {new_status}.'
            )

        except Invoice.DoesNotExist:
            messages.error(request, 'Invoice not found.')

    return redirect('user_dashboard')



# Helper function to normalize keys for Invoice Remarks to ensure consistent 
# storage and retrieval regardless of user input variations in the modal popup for remarks 
def _normalize_remark_key(value):
    text = (value or '').strip().lower()
    text = re.sub(r'[^a-z0-9]+', '_', text)
    return text.strip('_')


# E-Invoice Comparison View - Compare OCR extracted data with EInvoiceRegister database
@login_required
def get_einvoice_comparison(request, invoice_id):
    """
    Returns E-Invoice comparison data:
    - Uploaded Invoice data from raw_ocr_response
    - Govt Data from EInvoiceRegister table
    - Match status for each field
    """
    try:
        invoice = get_object_or_404(                               # fetch invoice by ID
            Invoice,                                               # ensures user can only access their company data
            id=invoice_id,
            company=request.user.company_code,
        )


        def _clean(value):                                    # these function normalize data before comparison
            if value is None:                                 # converts everything to string
                return ""
            if isinstance(value, datetime):                   # handles none and formats datetime
                return value.strftime("%d/%m/%Y")
            return str(value).strip()
   

        def _normalize_date(value):                           # normalize date to DD/MM/YYYY for better comparsion 
            text = _clean(value)
            if not text:
                return ""
            text = text.replace("-", "/").replace(".", "/")
            parts = text.split("/")
            if len(parts) == 3 and len(parts[0]) == 4:
                return f"{parts[2].zfill(2)}/{parts[1].zfill(2)}/{parts[0]}"
            return text

        def _normalize_amount(value):                         # removes commas, currency symbols and extra spaces for better comparison of amounts 
            text = _clean(value)
            if not text:
                return ""
            return text.replace(",", "").replace("₹", "").strip()


        def _normalize_gstin(value):                         # removes spaces/special characters and converts to uppercase                   
            text = _clean(value).upper()
            return re.sub(r"[^A-Z0-9]", "", text)

        ocr_data = invoice.raw_ocr_response or {}            
        invoice_data = ocr_data.get("result", {}).get("Invoice_data", {})      # reads raw ocr json and extracts important fields


        # Extract key fields from OCR
        ocr_fields = {
            "Supplier GSTIN": _clean(invoice_data.get("Vendor Gst No.")),      # 
            "Document Number": _clean(invoice_data.get("InvoiceId")),
            "Document Date": _clean(invoice_data.get("InvoiceDate")),
            "Invoice Amount": _clean(invoice_data.get("InvoiceTotal")),
            "IRN": _clean(invoice_data.get("Irn_No")),
            "IRN Status": "",
            "Ack No": "",
        }

        # Try to find matching E-Invoice Register record by IRN
        irn = _clean(ocr_fields.get("IRN"))
        supplier_gstin = _normalize_gstin(ocr_fields.get("Supplier GSTIN"))       
        document_number = _clean(ocr_fields.get("Document Number"))
        einvoice_record = None
        match_basis = "none"
        comparison_data = []

        if irn:
            try:
                einvoice_record = EInvoiceRegister.objects.get(irn=irn)
                match_basis = "irn"
            except EInvoiceRegister.DoesNotExist:
                einvoice_record = None

        # Fallback 1: Supplier GSTIN + Document Number
        if einvoice_record is None and supplier_gstin and document_number:
            gstin_candidates = EInvoiceRegister.objects.filter(document_number__iexact=document_number)
            for candidate in gstin_candidates:
                if _normalize_gstin(candidate.supplier_gstin) == supplier_gstin:
                    einvoice_record = candidate
                    match_basis = "gstin_document"
                    break

        # Fallback 2: Supplier GSTIN only (soft validation)
        if einvoice_record is None and supplier_gstin:
            gstin_candidates = EInvoiceRegister.objects.order_by("-irn_date")
            for candidate in gstin_candidates:
                if _normalize_gstin(candidate.supplier_gstin) == supplier_gstin:
                    einvoice_record = candidate
                    match_basis = "gstin"
                    break

        # Build comparison table
        fields_to_compare = [
            ("Supplier GSTIN", "supplier_gstin"),
            ("Document Number", "document_number"),
            ("Document Date", "document_date"),
            ("Invoice Amount", "amount"),
            ("IRN", "irn"),
            ("IRN Status", "irn_status"),
            ("Ack No", "ack_no"),
        ]

        for field_label, db_field in fields_to_compare:
            uploaded_value = _clean(ocr_fields.get(field_label))
            govt_value = ""
            match_status = "Mismatch"
            

            if einvoice_record:
                db_value = getattr(einvoice_record, db_field, "")
                govt_value = _clean(db_value)

                # Compare values
                uploaded_compare = uploaded_value.lower()
                govt_compare = govt_value.lower()

                if field_label == "Document Date":
                    uploaded_compare = _normalize_date(uploaded_value).lower()
                    govt_compare = _normalize_date(govt_value).lower()

                elif field_label == "Invoice Amount":
                    uploaded_compare = _normalize_amount(uploaded_value)
                    govt_compare = _normalize_amount(govt_value)


                if uploaded_compare == govt_compare and uploaded_compare:
                    match_status = "Matched"

                elif field_label == "Invoice Amount" and uploaded_value and govt_value:
                    # For amounts, allow small tolerance
                    try:
                        uploaded_amt = Decimal(_normalize_amount(uploaded_value))
                        govt_amt = Decimal(_normalize_amount(govt_value))
                        if abs(uploaded_amt - govt_amt) <= Decimal("1"):
                            match_status = "Matched"
                    except Exception:
                        pass


                elif field_label == "Document Date" and uploaded_value and govt_value:
                    # For dates, normalize formats
                    if _normalize_date(uploaded_value) == _normalize_date(govt_value):
                        match_status = "Matched"

                elif not uploaded_value or not govt_value:
                    #match_status = "Not Available"
                    match_status = "Missing"
            else:
                # No E-Invoice record found
                govt_value = "-"
                match_status = "Not Found"


            if field_label == "Supplier GSTIN" and supplier_gstin and match_basis == "gstin":
                # Soft signal for user: GSTIN exists in register even if other fields need review.
                #match_status = "Present in Register"
                match_status = "Matched"


            comparison_data.append({
                "field_name": field_label,
                "uploaded_value": uploaded_value or "-",
                "govt_value": govt_value or "-",
                "match_status": match_status,
            })


        if match_basis == "irn":
            simple_message = "Matched using IRN."

        elif match_basis == "gstin_document":
            simple_message = "Matched using Supplier GSTIN and Document Number."

        elif match_basis == "gstin":
            simple_message = "Supplier GSTIN is present in E-Invoice Register. Other fields are shown for review."

        else:
            simple_message = "No direct E-Invoice record found. Please review uploaded values."
        

        summary = {
            "matched": 0,
            "mismatch": 0,
            "missing": 0,
        }

        for row in comparison_data:
            if row["match_status"] == "Matched":
                summary["matched"] += 1
            elif row["match_status"] == "Mismatch":
                summary["mismatch"] += 1
            elif row["match_status"] == "Missing":
                summary["missing"] += 1

        return JsonResponse({
            "success": True,
            "record_found": einvoice_record is not None,
            "match_basis": match_basis,
            "simple_message": simple_message,
            "comparison_data": comparison_data,
            "summary": summary,
        })


    except Exception as e:
        print(f"E-Invoice comparison error: {str(e)}")
        return JsonResponse({
            "success": False,
            "error": str(e),
        }, status=500)



# Save Invoice Remark View for Processor Role to save Remarks per Invoice 
# This view will be called from frontend when processor adds remark for any validation paramater 
# in modal popup and we are saving those remarks in InvoiceRemark model with relation to invoice
# and also returning the saved remark in response to show updated remark in modal without refresh
@login_required
@require_http_methods(["POST"])
def save_invoice_remark(request):
    role = normalize_user_role(getattr(request.user, 'role', ''))
    if role != 'PROCESSOR':
        return JsonResponse({'error': 'Only Processor can save remarks.'}, status=403)

    try:
        payload = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON payload.'}, status=400)

    invoice_id = payload.get('invoice_id')
    section = _normalize_remark_key(payload.get('section'))
    subsection = _normalize_remark_key(payload.get('subsection'))
    parameter_key = _normalize_remark_key(payload.get('parameter_key'))
    remark_text = (payload.get('remark') or '').strip()

    if not all([invoice_id, section, subsection, parameter_key]):
        return JsonResponse({'error': 'Missing required fields.'}, status=400)

    invoice = get_object_or_404(
        Invoice,
        id=invoice_id,
        company=request.user.company_code,
    )

    remark_obj, created = InvoiceRemark.objects.get_or_create(
        invoice=invoice,
        section=section,
        subsection=subsection,
        parameter_key=parameter_key,
        defaults={
            'remark': remark_text,
            'created_by': request.user,
            'updated_by': request.user,
        },
    )

    if not created:
        remark_obj.remark = remark_text
        remark_obj.updated_by = request.user
        remark_obj.save(update_fields=['remark', 'updated_by', 'updated_at'])

    return JsonResponse({
        'message': 'Remark saved successfully.',
        'remark': {
            'invoice_id': invoice.id,
            'section': remark_obj.section,
            'subsection': remark_obj.subsection,
            'parameter_key': remark_obj.parameter_key,
            'remark': remark_obj.remark,
            'updated_at': remark_obj.updated_at,
        }
    })


@login_required
@require_http_methods(["GET"])
def get_invoice_remarks(request, invoice_id):
    invoice = get_object_or_404(
        Invoice,
        id=invoice_id,
        company=request.user.company_code,
    )

    section = _normalize_remark_key(request.GET.get('section'))
    remarks_qs = InvoiceRemark.objects.filter(invoice=invoice)
    if section:
        remarks_qs = remarks_qs.filter(section=section)

    remarks = [
        {
            'section': item.section,
            'subsection': item.subsection,
            'parameter_key': item.parameter_key,
            'remark': item.remark,
            'updated_at': item.updated_at,
        }
        for item in remarks_qs
    ]

    return JsonResponse({'invoice_id': invoice.id, 'remarks': remarks})


#-------------------------------------------------------------------------------------------------------------------------------------------

# -----======================== MODAL POPUP INVOICE BUTTON FUNCTIONALITY ========================-----
# In this view we will get the raw data from session and return as JsonResponse
# Get Invoice Raw Data View for Modal Popup
# On Pending Action tab then Invoice Status then when Invoice button 
# is clicked this view will be called to get the raw data from session 
# This is the main function that will be called when user clicks on Invoice button in Pending Action tab to show the raw data in modal popup
# and we are using the session data that we stored in update_invoice_status view when user clicks approve button to store raw data in session for that invoice and then here we are fetching that data to show in modal popup

# 2nd version of get_invoice_raw_data view
@login_required
def get_invoice_raw_data(request, invoice_id):

    """
    Returns invoice data for the modal popup.
    Currently using data from data_gathering.py
    """
 
    invoice = get_object_or_404(
        Invoice,
        id=invoice_id,
        company=request.user.company_code,
    )

    data = invoice.raw_ocr_response or {}
    invoice_meta = data.get("result", {}).get("Invoice_data", {})
    invoice_items = invoice_meta.get("Invoice items:", {}) or {}
    
    # ==================================== Invoice Calculation 1 Data Extraction ==================================== #

    table_check = data.get("result", {}).get("CHECKS", {}).get("table_data", {}).get("Table_Check_data", "[]")

    import json
    try:
        table_check_list = json.loads(table_check)
    except:
        table_check_list = []

    TABLE_LINE_TOLERANCE = Decimal("1")

    def to_decimal(val):
        if val is None:
            return None
        if isinstance(val, Decimal):
            return val
        if isinstance(val, (int, float)):
            return Decimal(str(val))
        if isinstance(val, str):
            cleaned = val.replace(",", "").replace("%", "").replace("₹", "").strip()
            if not cleaned:
                return None
            try:
                return Decimal(cleaned)
            except:
                return None
        return None


    def get_normalized_text(value):
        if value is None:
            return ""
        text = str(value).strip()
        if text in {"", "-", "--", "N/A", "n/a", "null", "None"}:
            return ""
        return text

    def pick_product_code(item, fallback_item=None):
        candidate_keys = [
            "product_code", "productCode", "Product_Code", "Product Code",
            "hsn_code", "HSN_Code", "HSN Code", "hsn", "HSN",
            "hsn_sac", "HSN/SAC", "HSN/SAC Code",
            "sac", "sac_code", "SAC", "SAC_Code",
            "item_hsn", "item_hsn_code", "item_code", "code",
        ]

        for key in candidate_keys:
            value = get_normalized_text(item.get(key))
            if value:
                return value

        if isinstance(fallback_item, dict):
            for key in candidate_keys:
                value = get_normalized_text(fallback_item.get(key))
                if value:
                    return value

        return ""

    def compute_line_fields(item, row_index):
        invoice_item_fallback = invoice_items.get(f"item#{row_index + 1}", {}) if isinstance(invoice_items, dict) else {}
        item["product_code"] = pick_product_code(item, invoice_item_fallback)

        qty = to_decimal(item.get("item_quantity"))
        unit = to_decimal(item.get("unit_price"))
        amount = to_decimal(item.get("amount"))
        tax_rate = to_decimal(item.get("tax_rate"))

        calc = None

        # ---------------- Basic Calculation ---------------- #
        if qty is not None and unit is not None:
            calc = (qty * unit).quantize(Decimal("0.01"))
            item["qty_unitprice"] = float(calc)
        else:
            item["qty_unitprice"] = "-"

        # ---------------- Smart Matching Logic ---------------- #
        if calc is not None and amount is not None:

            matched = False

            # Case 1: Amount = Basic only
            if abs(calc - amount) <= TABLE_LINE_TOLERANCE:
                matched = True

            # If tax exists, check tax scenarios
            if tax_rate is not None:

                tax_amount = (calc * tax_rate / Decimal("100")).quantize(Decimal("0.01"))

                # Case 2: Amount = Basic + 1 Tax (IGST)
                if abs((calc + tax_amount) - amount) <= TABLE_LINE_TOLERANCE:
                    matched = True

                # Case 3: Amount = Basic + 2 Tax (CGST + SGST)
                if tax_rate <= Decimal("14"):   # prevents false 2x for 18% IGST
                    if abs((calc + (tax_amount * 2)) - amount) <= TABLE_LINE_TOLERANCE:
                        matched = True

            item["check1"] = "Matched" if matched else "Not Matched"

        else:
            item["check1"] = "-"

        # ---------------- Extra Tax Display Columns ---------------- #
        if calc is not None and tax_rate is not None:
            tax_amount = (calc * tax_rate / Decimal("100")).quantize(Decimal("0.01"))

            item["qty_unit+rate_qty_unit"] = float(calc + tax_amount)
            item["qty_unit+2_rate_qty_unit"] = float(calc + (tax_amount * 2))
        else:
            item["qty_unit+rate_qty_unit"] = "-"
            item["qty_unit+2_rate_qty_unit"] = "-"

        return item


    # Apply calculation to each row
    for idx, item in enumerate(table_check_list):
        compute_line_fields(item, idx)


    # ---------------- Invoice Level Basic Amount Comparison ---------------- #

    table_sum_dec = Decimal("0")

    for item in table_check_list:
        line_total = to_decimal(item.get("qty_unitprice"))
        if line_total is not None:
            table_sum_dec += line_total

    ocr_amount_dec = to_decimal(invoice_meta.get("SubTotal"))

    if table_check_list and ocr_amount_dec is not None:
        diff = abs(ocr_amount_dec - table_sum_dec)
        status = "Matched" if diff <= TABLE_LINE_TOLERANCE else "Not Matched"
        ocr_amount_val = float(ocr_amount_dec)
        table_sum_val = float(table_sum_dec)
    else:
        status = "-"
        ocr_amount_val = "-"
        table_sum_val = "-"

    ic1_data = {
        "basic_amount": {
            "ocr_amount": ocr_amount_val,
            "table_sum": table_sum_val,
            "status": status
        }
    }

    # ==================================== Account Check Data Extraction ==================================== #

    account_check = data.get("result", {}).get("CHECKS", {}).get("Account_check", {})

    def extract_account_param(key):
        param = account_check.get(key, {})
        return {
            "gst_portal": param.get("Gst_Portal", ""),
            "invoice_data": param.get("Invoice_data", ""),
            "status": param.get("status", "")
        }
        
        
    account_check_full = {
        
        # not showing in frontend as per new BRD but keeping the code ready if we need to show it later 
        #"Complete_Invoice": extract_account_param("Complete_Invoice"),

        "Customer_Adress": extract_account_param("Customer_Adress"),
        "Customer_Name": extract_account_param("Customer_Name"),
        
        # not showing in frontend as per new BRD but keeping the code ready if we need to show it later
        #"Invoice_Blocked_Credit": extract_account_param("Invoice_Blocked_Credit"),

        "Invoice_Date": extract_account_param("Invoice_Date"),
        "Invoice_Number": extract_account_param("Invoice_Number"),

        # not showing in frontend as per new BRD but keeping the code ready if we need to show it later
        #"Invoice_RCM-Services": extract_account_param("Invoice_RCM-Services"),  

        "Pre_year": extract_account_param("Pre_year"),
        "gstnumber_gstcharged": extract_account_param("gstnumber_gstcharged"),
        "valid_invoice": extract_account_param("valid_invoice")
    }

    #--------------------------------------------------------------------------------------
    # Fallback for GST charged row when Account_check fields are blank.
    # showing Gst number Gst charged in frontend using full json data using fallback logic because it is not important that 
    # it can always come from Account_check, we can show it from other place if it is missing in Account_check and it is important 
    # field for user to see in frontend for better clarity and decision making. So added fallback logic to show it from other place
    # if it is missing in Account_check. We can do this for other fields as well in future if needed.
    gst_charged = account_check_full.get("gstnumber_gstcharged", {})
    if not gst_charged.get("gst_portal"):
        gst_charged["gst_portal"] = (
            data.get("result", {})
            .get("CHECKS", {})
            .get("data_from_gst", {})
            .get("vendor_gst_data", {})
            .get("Gstin", "")
        )

    if not gst_charged.get("invoice_data"):
        gst_charged["invoice_data"] = (
            account_check.get("Vendor_Gst_mentioned", {}).get("Invoice_data", "")
            or invoice_meta.get("Vendor Gst No.", "")
        )
    
    #--------------------------------------------------------------------------------------

    customer_name = account_check.get("Customer_Name", {})
    customer_address = account_check.get("Customer_Adress", {}) 

    gst_address_text = customer_address.get("Gst_Portal", "")
    invoice_address_text = customer_address.get("Invoice_data", "")

    account_check_data = {
        "customer_name": {
            "gst_portal": customer_name.get("Gst_Portal", "-"),
            "invoice_data": customer_name.get("Invoice_data", "-"),
            "status": customer_name.get("status", "-")
        },
        "customer_address": {
            "gst_portal": gst_address_text or "-",
            "invoice_data": invoice_address_text or "-",
            "status": customer_address.get("status", "-"),
            "match_percentage": customer_address.get("match_percentage", "-")
        }
    }

    other_parameter_check = {
        "invoice_date": extract_account_param("Invoice_Date"),
        "invoice_number": extract_account_param("Invoice_Number"),
        "previous_year": extract_account_param("Pre_year"),
        "gst_mentioned": extract_account_param("gstnumber_gstcharged"),
        "valid_invoice": extract_account_param("valid_invoice")
    }

    # ==================================== Invoice Calculation 2 Data Extraction ==================================== #

    INVOICE_CALCULATION_2_TOLERANCE = 1

    table_sum_f = float(table_sum_val) if table_sum_val != "-" else None

    ocr_tax = invoice_meta.get("TotalTax")
    ocr_total = invoice_meta.get("InvoiceTotal")

    def to_float(val):
        try:
            return float(val)
        except:
            return None

    ocr_tax_f = to_float(ocr_tax)
    ocr_total_f = to_float(ocr_total)

    if table_sum_f is not None and ocr_tax_f is not None and ocr_total_f is not None:
        expected_total = table_sum_f + ocr_tax_f
        diff = abs(expected_total - ocr_total_f)
        ic2_status = "Matched" if diff <= INVOICE_CALCULATION_2_TOLERANCE else "Not Matched"
    else:
        expected_total = "-"
        ic2_status = "-"

    invoice_calculation_2 = {
        "total_amount": {
            "table_sum": table_sum_f if table_sum_f is not None else "-",
            "ocr_tax": ocr_tax_f if ocr_tax_f is not None else "-",
            "ocr_total": ocr_total_f if ocr_total_f is not None else "-",
            "status": ic2_status
        }
    }

    # ============================= Tax Should Be Charged (HSN/SAC Lookup) =========================== #

    selected_product_code = ""
    for item in table_check_list:
        if not isinstance(item, dict):
            continue
        candidate_code = get_normalized_text(item.get("product_code"))
        if candidate_code:
            selected_product_code = candidate_code
            break

    hsn_match = HSN.objects.filter(hsn_code=selected_product_code).first() if selected_product_code else None
    sac_match = SAC.objects.filter(sac_code=selected_product_code).first() if selected_product_code else None

    tax_should_be_charged = {
        "product_code": selected_product_code or None,
        "tax_rate": None,
        "block_credit": None,
        "rcm": None,
    }

    if hsn_match:
        tax_should_be_charged.update({
            "tax_rate": hsn_match.tax_rate,
            "block_credit": hsn_match.block_credit,
            "rcm": hsn_match.rcm,
        })
    elif sac_match:
        tax_should_be_charged.update({
            "tax_rate": sac_match.tax_rate,
            "block_credit": sac_match.block_credit,
            "rcm": sac_match.rcm,
        })

    # ============================= PAN CHECK =========================== #

    tax_check_raw = data.get("result", {}).get("CHECKS", {}).get("tax_check", {})

    pan_check_data = {
        "vendor_206ab": tax_check_raw.get("Vendor_206AB", {}).get("status", "-"),
        "vendor_pan_active": tax_check_raw.get("Vendor_Pan_Active", {}).get("status", "-"),
        "vendor_pan_aadhar_linked": tax_check_raw.get("Vendor_Pan-Adhar_Linked", {}).get("status", "-"),
    }

    # ============================= Tax Check =========================== #

    tax_items = invoice_meta.get("Tax Items", {}) or {}

    def safe_amount(tax_key):
        tax_obj = tax_items.get(tax_key, {}) or {}
        return to_float(tax_obj.get("amount")) or 0

    igst_amount = safe_amount("IGST")
    cgst_amount = safe_amount("CGST")
    sgst_amount = safe_amount("SGST")

    if igst_amount > 0:
        tax_type = "IGST"
    elif cgst_amount > 0 or sgst_amount > 0:
        tax_type = "CGST + SGST"
    else:
        tax_type = "-"

    try:
        subtotal = to_float(invoice_meta.get("SubTotal")) or 0
        total_tax = to_float(invoice_meta.get("TotalTax")) or 0
        tax_rate_percent = round((total_tax / subtotal) * 100, 2) if subtotal > 0 else 0
        tax_rate = f"{tax_rate_percent}%"
    except:
        tax_rate = "-"

    tax_check_data = {
        "tax_rate": tax_rate,
        "tax_type": tax_type,
        "input_credit": account_check.get("Invoice_Blocked_Credit", {}).get("status", "-"),
        "rcm_fc": account_check.get("Invoice_RCM-Services", {}).get("status", "-"),
    }

    response_data = {
        "invoice_calculation_1": ic1_data["basic_amount"],
        "invoice_calculation_2": invoice_calculation_2["total_amount"],
        "account_check": account_check_data,
        "account_check_full": account_check_full,
        "tax_check": tax_check_data,
        "tax_should_be_charged": tax_should_be_charged,
        "pan_check": pan_check_data,
        "other_parameter_check": other_parameter_check,
        "table_check": table_check_list,
        "result": data.get("result", {})
    }

    return JsonResponse(response_data, safe=False)



# Tax Check API View (for Tax button in modal popup, separate API for better modularity and performance)
from rest_framework.decorators import api_view
from rest_framework.response import Response
import json
@api_view(["POST"])
def tax_check_view(request):

    data = request.data or {}
    
    # STEP 1: Extract Table_Check_data
    table_data = data.get("table_data", {})
    table_check_string = table_data.get("Table_Check_data")

    # Convert string → list
    try:
        table_items = json.loads(table_check_string)
    except:
        return Response({"error": "Invalid Table_Check_data"})

    # ✅ STEP 2: Check empty
    if not table_items:
        return Response({"error": "No items found"})

    # ✅ STEP 3: Take ONLY first item
    first_item = table_items[0]
    product_code = first_item.get("product_code")

    if not product_code:
        return Response({"error": "Product code not found"})

    product_code = str(product_code).strip()

    # ✅ STEP 4: Search in HSN
    hsn = HSN.objects.filter(hsn_code=product_code).first()

    if hsn:
        return Response({
            "tax_should_be_charged": {
                "product_code": product_code,
                "tax_rate": hsn.tax_rate,
                "block_credit": hsn.block_credit,
                "rcm": hsn.rcm
            }
        })

    # ✅ STEP 5: Search in SAC
    sac = SAC.objects.filter(sac_code=product_code).first()

    if sac:
        return Response({
            "tax_should_be_charged": {
                "product_code": product_code,
                "tax_rate": sac.tax_rate,
                "block_credit": sac.block_credit,
                "rcm": sac.rcm
            }
        })

    # ✅ STEP 6: Not found
    return Response({
        "tax_should_be_charged": {
            "product_code": product_code,
            "tax_rate": None,
            "block_credit": None,
            "rcm": None
        }
    })


    
# Download Validation Summary View
# This view generates and serves an Excel file summarizing the validation checks for a given invoice
# 2nd version with dynamic summary data population and improved formatting 
@login_required
def download_validation_summary(request, invoice_id):
    
    # Download validation summary for an invoice in Excel format
    try:
        invoice = Invoice.objects.get(
            id=invoice_id,
            company=request.user.company_code
        )

    except Invoice.DoesNotExist:
        messages.error(request, 'Invoice not found.')
        return redirect('user_dashboard')


    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Summary"


    # Header styling
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)


    # Title
    ws['A1'] = 'Invoice Validation Summary'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')


    # Invoice Details Section
    ws['A3'] = 'Invoice Details'
    ws['A3'].font = Font(bold=True, size=14)

    details = [
        ['Vendor Name', invoice.vendor_name],
        ['Invoice Number', invoice.invoice_number],
        ['Invoice Date', invoice.invoice_date.strftime('%d %b %Y')],
        ['Invoice Value', f'₹{invoice.invoice_value}'],
        ['Status', invoice.status],
        ['Upload Date', invoice.upload_date.strftime('%d %b %Y %H:%M')],
        ['Uploaded By', invoice.uploaded_by.username if invoice.uploaded_by else 'N/A'],
        ['File Name', invoice.file_name],
        ['Response', invoice.response or 'N/A'],
    ]

    row = 4
    for label, value in details:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

   
    # Validation Checks Section       
    row += 2
    ws[f'A{row}'] = 'Validation Checks'
    ws[f'A{row}'].font = Font(bold=True, size=14)

    row += 1

    check_headers = ['Check Type', 'Status', 'Details', 'Result']
    for col_idx, header in enumerate(check_headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')



    # Fetch Validation Summary
 

    import json
    summary = invoice.validation_summary

    # If stored as string , convert to dict
    if isinstance(summary, str):
        try: 
            summary = json.loads(summary)
        except:
            summary = {}

    if not summary:
        summary = {"categories": []}
    
    row += 1

    
    # 2nd version with dynamic summary data population and improved formatting 
    # Populate Validation Rows

    summary = invoice.validation_summary
    
    # If stored as string, convert to dict
    if isinstance(summary, str):
        try:
            summary = json.loads(summary)
        except:
            summary = {}
    
    if not summary:
        summary = {"categories": []}
    
    row += 1
    
    # Iterate through categories and checks
    for category_data in summary.get("categories", []):
        category_name = category_data.get("category", "")
        checks = category_data.get("checks", [])
        
        for check in checks:
            check_type = check.get("check_name", "")
            status = check.get("status", "")
            details = check.get("details", "")
            
            # Determine result based on status
            if status == "Pass":
                result = "Valid"
                status_color = "008000"  # Green
            elif status == "Fail":
                result = "Invalid"
                status_color = "FF0000"  # Red
            else:
                result = "Pending"
                status_color = "FFA500"  # Orange
            
            # Column 1 - Check Type
            ws.cell(row=row, column=1, value=check_type)
            
            # Column 2 - Status
            status_cell = ws.cell(row=row, column=2, value=status)
            status_cell.font = Font(color=status_color, bold=True)
            
            # Column 3 - Details
            ws.cell(row=row, column=3, value=details)
            
            # Column 4 - Result
            result_cell = ws.cell(row=row, column=4, value=result)
            result_cell.font = Font(color=status_color, bold=True)
            
            row += 1


    # Column Widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 15

    # Prepare Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="Invoice_{invoice.invoice_number}_Validation_Summary.xlsx"'
    )

    wb.save(response)
    return response


#---------------------------------------------------------------------------------------------------
# Edits an existing GST registration record for the logged-in user's company.
# - Ensures only authenticated users can access this view.
# - Restricts editing to GST records belonging to the user's company.
# - Processes form submission via POST request.
# - Displays success or validation error messages accordingly.
# - Redirects back to the dashboard Settings section after processing.
@login_required
def edit_gst_detail(request, gst_id):
    """
    Edit existing GST registration detail.
    Only allows editing GST details belonging to the user's company.
    """
    gst_detail = get_object_or_404(GSTDetails, id=gst_id, company=request.user.company_code)
    
    if request.method == 'POST':
        form = GSTDetailsForm(request.POST, instance=gst_detail)
        if form.is_valid():
            form.save()
            messages.success(request, f'GST Details for {gst_detail.state} updated successfully!')
            response = redirect('company_admin_dashboard')
            response['Location'] = response['Location'] + '?section=settings'
            return response
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    
    response = redirect('company_admin_dashboard')
    response['Location'] = response['Location'] + '?section=settings'
    return response
#--------------------------------------------------------------------------------------------------


#-------------------------------------------------------------------------------------------------
# This View Handles Deletion of GST Details
# Deletes a GST registration record for the logged-in user's company.
# - Ensures only authenticated users can access this view.
# - Restricts deletion to GST records belonging to the user's company.
# - Allows deletion only via POST request for safety.
# - After deletion, redirects back to the dashboard Settings section
#   with a success message.
@login_required
def delete_gst_detail(request, gst_id):
    """
    Delete GST registration detail.
    Only allows deleting GST details belonging to the user's company.
    """
    gst_detail = get_object_or_404(GSTDetails, id=gst_id, company=request.user.company_code)
    
    if request.method == 'POST':
        state_name = gst_detail.state
        gst_detail.delete()
        messages.success(request, f'GST Details for {state_name} deleted successfully!')
    
    response = redirect('company_admin_dashboard')
    response['Location'] = response['Location'] + '?section=settings'
    return response
#-------------------------------------------------------------------------------------------------

